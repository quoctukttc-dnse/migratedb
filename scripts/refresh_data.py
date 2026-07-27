#!/usr/bin/env python3
"""
Refresh the SO Migration dashboard's embedded data from the source Excel file.

Usage:
    python3 scripts/refresh_data.py <path-to-xlsx> <mtime-ms> <repo-root>

Reads the 3 relevant sheets from the "Danh Sach SO Can Migration.xlsx" workbook,
canonicalizes RCM names, computes totals + per-RCM aggregates, and replaces the
`const DATA = {...};` block inside <repo-root>/index.html plus
<repo-root>/data/dashboard_data.json. Does NOT touch styling/layout/markup —
only the embedded data payload.

Exits 0 with "NO_CHANGE" printed if the newly computed payload is byte-identical
(ignoring report_date) to what's already in the repo, so the caller can skip an
empty commit. Exits 0 with "UPDATED" printed otherwise.
"""
import sys, json, collections, unicodedata, re
from datetime import datetime, timezone, timedelta

def clean(v):
    if v is None: return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v

def strip_diacritics(s):
    if s is None: return ''
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd').replace('Đ', 'D')
    return s.lower().strip()

def norm_words(s):
    return set(re.findall(r'[a-z]+', strip_diacritics(s)))

def norm_cust_key(c):
    if c is None: return None
    s = str(c).strip()
    if not s: return None
    s = re.sub(r'[.,]', '', s)
    s = re.sub(r'\s+', ' ', s).upper().strip()
    return s or None

CANONICAL = [
    'TRẦN NGỌC THÁI HÀ', 'TRẦN HỮU CÁT TƯỜNG', 'LÊ THÚY HẰNG', 'NGUYỄN LÊ TRƯỜNG AN',
    'TRẦN THỊ THANH HÀ', 'PHẠM THỊ THU TRANG', 'LÊ THỊ THẢO NGUYÊN', 'TRẦN THỊ THANH THẢO',
    'MAI THÙY DUNG', 'QUÁCH BẢO NGHĨA', 'TRẦN THỊ KIỀU MY', 'THIỆU MỸ LINH', 'TRẦN ANH QUỐC',
    'LÊ CẨM GIAI KỲ', 'NGUYỄN THỊ THU HƯƠNG',
]
_cache = {}
def canon(v):
    v = clean(v)
    if v is None:
        return 'Chưa gán'
    if v in _cache:
        return _cache[v]
    vw = norm_words(v)
    result = v
    for c in CANONICAL:
        if vw and vw.issubset(norm_words(c)):
            result = c
            break
    _cache[v] = result
    return result

def norm_so(v):
    if v is None:
        return None
    return str(v).strip() or None


def pct_of(part, base):
    return round(100 * part / base, 1) if base else 0


PRIORITY_BUCKETS = ['1', '2', '3', 'x', 'blank']


def pbucket(v):
    """Bucket a raw '2. SCAF SO Detail' Priority cell into '1'/'2'/'3'/'x'/'blank'."""
    if v is None:
        return 'blank'
    if v == 1 or v == '1':
        return '1'
    if v == 2 or v == '2':
        return '2'
    if v == 3 or v == '3':
        return '3'
    if isinstance(v, str) and v.strip().lower() == 'x':
        return 'x'
    return 'blank'


def main():
    import openpyxl
    xlsx_path, mtime_ms, repo_root = sys.argv[1], int(sys.argv[2]), sys.argv[3]
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Load '2. SCAF SO Detail' early so we can cross-check the "Upload to SAP"
    # flag from '1. SO Monitor Migration' against it (see note below).
    ws3 = wb['2. SCAF SO Detail']
    scaf_rows = list(ws3.iter_rows(min_row=2, values_only=True))
    # SO_No_ScaX (col index 2 in SCAF SO Detail) is the ORIGINAL SCAX SO number —
    # the same value as 'SCAX SO' (col index 1) in '1. SO Monitor Migration'.
    # (col index 1 in SCAF SO Detail, "SO_No", is a *different*, SCAF-side
    # numbering — joining on it instead silently produces near-zero overlap.)
    # For each SCAX SO, note whether ANY of its SCAF Detail lines is flagged
    # in_SAP == 1.
    scax_has_sap = collections.defaultdict(bool)
    for r in scaf_rows:
        so_scax = norm_so(r[2])
        if so_scax and r[19] == 1:
            scax_has_sap[so_scax] = True

    # For each SCAX SO, the set of Priority buckets present among its SCAF
    # Detail lines (a single SO can span multiple lines with different
    # priorities). Used to filter '1. SO Monitor Migration' rows by priority
    # for the RCM/Customer detail-table filter — joined the same way as
    # scax_has_sap above (SO_No_ScaX, NOT SCAF Detail's own "SO_No").
    so_priorities = collections.defaultdict(set)
    for r in scaf_rows:
        so_scax = norm_so(r[2])
        if so_scax:
            so_priorities[so_scax].add(pbucket(r[21]))

    ws = wb['1. SO Monitor Migration']
    mon_rows = list(ws.iter_rows(min_row=2, values_only=True))

    def build_monitor_aggregates(priority_filter=None):
        """Aggregate '1. SO Monitor Migration' rows by RCM and by Customer.
        If priority_filter is set (one of PRIORITY_BUCKETS), only rows whose
        SO has at least one SCAF Detail line in that priority bucket are
        included (inclusive — a multi-line SO can match more than one
        bucket)."""
        by_rcm_ = collections.defaultdict(lambda: {
            'total_lines': 0, 'migrated_scaf': 0, 'not_migrated_scaf': 0, 'uploaded_sap': 0,
            'order_qty': 0.0, 'so_set': set(), 'customers': set(),
        })
        by_cust_ = collections.defaultdict(lambda: {
            'total_lines': 0, 'migrated_scaf': 0, 'not_migrated_scaf': 0, 'uploaded_sap': 0,
            'so_set': set(), 'variants': collections.Counter(),
        })
        for r in mon_rows:
            if priority_filter is not None:
                if priority_filter not in so_priorities.get(norm_so(r[1]), ()):
                    continue
            rcm = canon(r[5]); d = by_rcm_[rcm]
            d['total_lines'] += 1
            if isinstance(r[18], (int, float)):
                d['order_qty'] += r[18]
            is_scaf = clean(r[33]) == 'x'
            # "Upload to SAP" in Monitor is a manually-ticked 'x' column and is
            # heavily stale/under-marked (verified: only 398/9676 rows marked,
            # vs 3504/9676 once cross-checked against SCAF SO Detail's live
            # in_SAP flag for the same SCAX SO — a >8x undercount). We treat a
            # row as uploaded if EITHER the manual 'x' is present OR any SCAF
            # Detail line for that same SO is flagged in_SAP == 1.
            is_sap = clean(r[34]) == 'x' or scax_has_sap.get(norm_so(r[1]), False)
            if is_scaf:
                d['migrated_scaf'] += 1
            else:
                d['not_migrated_scaf'] += 1
            if is_sap:
                d['uploaded_sap'] += 1
            d['so_set'].add(r[1])
            if r[6]:
                d['customers'].add(r[6])

            key = norm_cust_key(r[6]) or 'CHƯA XÁC ĐỊNH'
            dc = by_cust_[key]
            dc['variants'][str(clean(r[6])) if clean(r[6]) is not None else 'Chưa xác định'] += 1
            dc['total_lines'] += 1
            if is_scaf:
                dc['migrated_scaf'] += 1
            else:
                dc['not_migrated_scaf'] += 1
            if is_sap:
                dc['uploaded_sap'] += 1
            dc['so_set'].add(r[1])
        return by_rcm_, by_cust_

    by_rcm, by_cust = build_monitor_aggregates(None)

    ws2 = wb['3. SO Khong Migration']
    km_rows = list(ws2.iter_rows(min_row=2, values_only=True))
    by_rcm_km = collections.defaultdict(lambda: {'total_lines': 0, 'approved': 0, 'so_set': set()})
    for r in km_rows:
        rcm = canon(r[5]); d = by_rcm_km[rcm]
        d['total_lines'] += 1
        if clean(r[34]) == 'x':
            d['approved'] += 1
        d['so_set'].add(r[1])

    # 'Reporting' sheet, block "1. Ưu Tiên 1 BY SO" (cols G..M, rows 4..18, 15 RCM rows).
    # This block is driven by live COUNTIFS formulas against '2. SCAF SO Detail'
    # (Priority == "1", is_PRI, is BOM, in SAP). We read the sheet's own cached
    # per-RCM formula results directly (data_only=True) rather than re-deriving
    # them, since that's exactly what the business sees when they open the workbook.
    #
    # NOTE: the sheet's own "Total" row (row 19) uses =SUM(I3:I17) etc., which is
    # off by one versus the actual data range (rows 4-18) — it includes the blank
    # header row 3 (contributes 0) but excludes the LAST RCM row (18, currently
    # TRẦN THỊ THANH THẢO), so the sheet's own Total undercounts by that RCM's
    # values every time. We deliberately do NOT use that cached Total cell;
    # we always recompute the total by summing the (verified-correct) per-RCM
    # rows ourselves so this dashboard's total is accurate even though the
    # source sheet's own Total cell is not.
    pri1_rows = []
    if 'Reporting' in wb.sheetnames:
        wsr = wb['Reporting']
        for r in wsr.iter_rows(min_row=4, max_row=40, min_col=7, max_col=13, values_only=True):
            rcm_raw = clean(r[0])
            if rcm_raw is None:
                continue
            if str(rcm_raw).strip() == 'Total':
                break
            so_ut1, pri, bom, sap = r[2] or 0, r[3] or 0, r[4] or 0, r[5] or 0
            pri1_rows.append({'rcm': canon(rcm_raw), 'so_ut1': so_ut1, 'pri': pri, 'bom': bom, 'sap': sap})
        # merge any duplicate canonical RCM rows (defensive, shouldn't normally happen)
        merged = collections.OrderedDict()
        for row in pri1_rows:
            m = merged.setdefault(row['rcm'], {'rcm': row['rcm'], 'so_ut1': 0, 'pri': 0, 'bom': 0, 'sap': 0})
            for k in ('so_ut1', 'pri', 'bom', 'sap'):
                m[k] += row[k]
        pri1_rows = list(merged.values())
        pri1_total = {k: sum(row[k] for row in pri1_rows) for k in ('so_ut1', 'pri', 'bom', 'sap')} if pri1_rows else None
    else:
        pri1_total = None

    # (scaf_rows already loaded near the top of main(), used there to build scax_has_sap)
    def build_scaf_rcm_aggregates(priority_filter=None):
        """Aggregate '2. SCAF SO Detail' rows by RCM. If priority_filter is
        set, only rows whose OWN Priority cell buckets to that value are
        included — precise, since each SCAF Detail line has exactly one
        Priority (unlike the SO-level join used for the Monitor sheet)."""
        by_rcm_scaf_ = collections.defaultdict(lambda: {
            'total': 0, 'is_pri': 0, 'is_bom': 0, 'in_sap': 0, 'complete': 0, 'so_set': set(),
        })
        for r in scaf_rows:
            if priority_filter is not None and pbucket(r[21]) != priority_filter:
                continue
            rcm = canon(r[5]); d = by_rcm_scaf_[rcm]
            d['total'] += 1
            if r[17] == 1: d['is_pri'] += 1
            if r[18] == 1: d['is_bom'] += 1
            if r[19] == 1: d['in_sap'] += 1
            if clean(r[23]) == 'COMPLETE': d['complete'] += 1
            d['so_set'].add(r[1])
        return by_rcm_scaf_

    by_rcm_scaf = build_scaf_rcm_aggregates(None)

    all_rcms = sorted(set(by_rcm) | set(by_rcm_km) | set(by_rcm_scaf))
    result = {}
    for rcm in all_rcms:
        m = by_rcm.get(rcm, {}); k = by_rcm_km.get(rcm, {}); s = by_rcm_scaf.get(rcm, {})
        result[rcm] = {
            'monitor_lines': m.get('total_lines', 0),
            'migrated_scaf': m.get('migrated_scaf', 0),
            'not_migrated_scaf': m.get('not_migrated_scaf', 0),
            'uploaded_sap': m.get('uploaded_sap', 0),
            'order_qty': round(m.get('order_qty', 0)),
            'so_count': len(m.get('so_set', set())),
            'customer_count': len(m.get('customers', set())),
            'km_lines': k.get('total_lines', 0),
            'km_approved': k.get('approved', 0),
            'scaf_total': s.get('total', 0),
            'scaf_is_pri': s.get('is_pri', 0),
            'scaf_is_bom': s.get('is_bom', 0),
            'scaf_in_sap': s.get('in_sap', 0),
            'scaf_complete': s.get('complete', 0),
        }

    totals = {k: sum(v[k] for v in result.values()) for k in
              ['monitor_lines', 'migrated_scaf', 'not_migrated_scaf', 'uploaded_sap', 'so_count',
               'km_lines', 'km_approved', 'scaf_total', 'scaf_is_pri', 'scaf_is_bom',
               'scaf_in_sap', 'scaf_complete']}

    dt = datetime.fromtimestamp(mtime_ms / 1000, tz=timezone(timedelta(hours=7)))
    report_date = dt.strftime('%d/%m/%Y %H:%M')

    rows = []
    for rcm, d in result.items():
        ml = d['monitor_lines']
        pct_scaf = round(100 * d['migrated_scaf'] / ml, 1) if ml else 0
        pct_sap = round(100 * d['uploaded_sap'] / ml, 1) if ml else 0
        st = d['scaf_total']
        pct_pri = round(100 * d['scaf_is_pri'] / st, 1) if st else 0
        pct_bom = round(100 * d['scaf_is_bom'] / st, 1) if st else 0
        pct_insap = round(100 * d['scaf_in_sap'] / st, 1) if st else 0
        rows.append({'rcm': rcm, **d, 'pct_scaf': pct_scaf, 'pct_sap': pct_sap,
                      'pct_pri': pct_pri, 'pct_bom': pct_bom, 'pct_insap': pct_insap})
    rows.sort(key=lambda r: -r['not_migrated_scaf'])

    pct = {
        'scaf': round(100 * totals['migrated_scaf'] / totals['monitor_lines'], 1),
        'sap': round(100 * totals['uploaded_sap'] / totals['monitor_lines'], 1),
        'pri': round(100 * totals['scaf_is_pri'] / totals['scaf_total'], 1),
        'bom': round(100 * totals['scaf_is_bom'] / totals['scaf_total'], 1),
        'insap': round(100 * totals['scaf_in_sap'] / totals['scaf_total'], 1),
    }

    cust_rows = []
    for key, d in by_cust.items():
        ml = d['total_lines']
        label = d['variants'].most_common(1)[0][0] if key != 'CHƯA XÁC ĐỊNH' else 'Chưa xác định'
        pct_scaf = round(100 * d['migrated_scaf'] / ml, 1) if ml else 0
        pct_sap = round(100 * d['uploaded_sap'] / ml, 1) if ml else 0
        cust_rows.append({
            'customer': label,
            'monitor_lines': ml,
            'migrated_scaf': d['migrated_scaf'],
            'not_migrated_scaf': d['not_migrated_scaf'],
            'uploaded_sap': d['uploaded_sap'],
            'so_count': len(d['so_set']),
            'pct_scaf': pct_scaf,
            'pct_sap': pct_sap,
        })
    cust_rows.sort(key=lambda r: -r['monitor_lines'])

    priority1 = None
    if pri1_total is not None:
        pri1_out_rows = []
        for row in pri1_rows:
            base = row['so_ut1']
            pri1_out_rows.append({
                **row,
                'pct_pri': pct_of(row['pri'], base),
                'pct_bom': pct_of(row['bom'], base),
                'pct_sap': pct_of(row['sap'], base),
            })
        pri1_out_rows.sort(key=lambda r: -r['so_ut1'])
        priority1 = {
            'total': {
                **pri1_total,
                'pct_pri': pct_of(pri1_total['pri'], pri1_total['so_ut1']),
                'pct_bom': pct_of(pri1_total['bom'], pri1_total['so_ut1']),
                'pct_sap': pct_of(pri1_total['sap'], pri1_total['so_ut1']),
            },
            'rcm': pri1_out_rows,
        }

    # Priority filter (1/2/3/x/blank) for the RCM and Customer detail tables.
    # Priority itself only exists as a reliable field on '2. SCAF SO Detail'
    # ('1. SO Monitor Migration' has a same-named column, but it's populated
    # with product-segment text like "PANTY"/"OUTERWEAR" for most rows, not
    # a priority tier — not usable). For each bucket we recompute:
    #  - Monitor-derived fields (monitor_lines, migrated_scaf, uploaded_sap,
    #    so_count, ...) by including only Monitor rows whose SO has at least
    #    one SCAF Detail line in that bucket (SO-level join, inclusive).
    #  - SCAF-derived fields (scaf_total, pct_pri, pct_bom, ...) by directly
    #    filtering SCAF Detail rows to that bucket's own Priority value
    #    (precise, line-level — no join needed since RCM naming already
    #    matches between the two sheets via canon()).
    def build_rcm_rows(by_rcm_x, by_rcm_scaf_x):
        all_rcms_x = sorted(set(by_rcm_x) | set(by_rcm_scaf_x))
        out = []
        for rcm in all_rcms_x:
            m = by_rcm_x.get(rcm, {}); s = by_rcm_scaf_x.get(rcm, {})
            ml = m.get('total_lines', 0)
            migrated_scaf = m.get('migrated_scaf', 0)
            uploaded_sap = m.get('uploaded_sap', 0)
            st = s.get('total', 0)
            scaf_is_pri = s.get('is_pri', 0)
            scaf_is_bom = s.get('is_bom', 0)
            out.append({
                'rcm': rcm,
                'monitor_lines': ml,
                'migrated_scaf': migrated_scaf,
                'not_migrated_scaf': m.get('not_migrated_scaf', 0),
                'uploaded_sap': uploaded_sap,
                'so_count': len(m.get('so_set', set())),
                'scaf_total': st,
                'scaf_is_pri': scaf_is_pri,
                'scaf_is_bom': scaf_is_bom,
                'pct_scaf': pct_of(migrated_scaf, ml),
                'pct_sap': pct_of(uploaded_sap, ml),
                'pct_pri': pct_of(scaf_is_pri, st),
                'pct_bom': pct_of(scaf_is_bom, st),
            })
        out.sort(key=lambda r: -r['monitor_lines'])
        return out

    def build_cust_rows(by_cust_x):
        out = []
        for key, d in by_cust_x.items():
            ml = d['total_lines']
            label = d['variants'].most_common(1)[0][0] if key != 'CHƯA XÁC ĐỊNH' else 'Chưa xác định'
            out.append({
                'customer': label,
                'monitor_lines': ml,
                'migrated_scaf': d['migrated_scaf'],
                'not_migrated_scaf': d['not_migrated_scaf'],
                'uploaded_sap': d['uploaded_sap'],
                'so_count': len(d['so_set']),
                'pct_scaf': pct_of(d['migrated_scaf'], ml),
                'pct_sap': pct_of(d['uploaded_sap'], ml),
            })
        out.sort(key=lambda r: -r['monitor_lines'])
        return out

    rcm_by_priority = {}
    customer_by_priority = {}
    for b in PRIORITY_BUCKETS:
        by_rcm_b, by_cust_b = build_monitor_aggregates(b)
        by_rcm_scaf_b = build_scaf_rcm_aggregates(b)
        rcm_by_priority[b] = build_rcm_rows(by_rcm_b, by_rcm_scaf_b)
        customer_by_priority[b] = build_cust_rows(by_cust_b)

    payload = {'report_date': report_date, 'totals': totals, 'pct': pct, 'rcm': rows, 'customer': cust_rows,
               'priority1': priority1, 'priority_buckets': PRIORITY_BUCKETS,
               'rcm_by_priority': rcm_by_priority, 'customer_by_priority': customer_by_priority}

    # Compare against what's currently committed (ignoring report_date) to decide if anything changed.
    data_path = f'{repo_root}/data/dashboard_data.json'
    changed = True
    try:
        with open(data_path, encoding='utf-8') as f:
            prev = json.load(f)
        prev_cmp = {k: v for k, v in prev.items() if k != 'report_date'}
        new_cmp = {k: v for k, v in payload.items() if k != 'report_date'}
        changed = prev_cmp != new_cmp
    except FileNotFoundError:
        changed = True

    data_json = json.dumps(payload, ensure_ascii=False)
    html_path = f'{repo_root}/index.html'
    with open(html_path, encoding='utf-8') as f:
        html = f.read()
    new_html, n = re.subn(r'const DATA = \{.*?\};', 'const DATA = ' + data_json.replace('\\', '\\\\') + ';',
                           html, flags=re.S)
    assert n == 1, f'expected 1 DATA block, found {n}'
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(new_html)
    with open(data_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print('UPDATED' if changed else 'NO_CHANGE', report_date)

if __name__ == '__main__':
    main()
